from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slugify(text: str) -> str:
    safe = []
    for ch in text.lower():
        if ch.isalnum():
            safe.append(ch)
        elif ch in {" ", "-", "_", "/"}:
            safe.append("-")
    slug = "".join(safe).strip("-")
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug or "item"


def find_header_row(sheet) -> int:
    for row_index in range(1, min(sheet.max_row, 10) + 1):
        values = [cell_text(sheet.cell(row_index, col).value) for col in range(1, sheet.max_column + 1)]
        joined = "|".join(values)
        if "材料标题" in joined and "跳转链接" in joined:
            return row_index
    raise ValueError("Could not find header row")


def normalize_headers(headers: list[str]) -> dict[str, int]:
    return {header: index for index, header in enumerate(headers)}


def infer_summary(role: str, title: str, source: str) -> str:
    if "HR企业级AI提效实践" in title:
        return "结合亿欧公开文章中的案例，展示腾讯 HR 如何借助企业级 AI 将原本依赖反复打磨的工作流程压缩到更高效的完成方式，适合用于 HR 场景宣导与价值说明。"
    if "重新定义人力资源工作方式" in title:
        return "从腾讯内部HR工作的日常流程出发，介绍CodeBuddy与WorkBuddy如何覆盖七大HR角色，赋能简历筛选评估、JD撰写、数据看板一键生成等30+场景。"
    if "HR Claw" in title:
        return "腾讯内部面向HR场景的应用入口或工具集合型知识库。"
    if "腾讯内部" in source:
        return f"面向 {role} 团队的内部参考材料，适合在培训或宣导后作为延伸阅读与实操入口。"
    return f"面向 {role} 场景的参考材料，适合在宣导或培训中帮助团队快速理解相关能力与使用方式。"


# 明确属于"岗位"维度的关键词；其余归入"场景"
ROLE_KEYWORDS = {"HR", "财务", "法务", "税务", "人力", "运营", "市场", "销售", "采购", "行政", "研发", "产品", "设计", "客服", "IT"}


def classify_dimension(label: str) -> str:
    """Return 'role' if label matches a job function, else 'scene'."""
    for kw in ROLE_KEYWORDS:
        if kw in label:
            return "role"
    return "scene"


def infer_material_type(link: str, source: str) -> str:
    if "mp.weixin.qq.com" in link:
        return "文章"
    if "app-market" in link:
        return "工具入口"
    if "hrainative" in link:
        return "专题页面"
    if "腾讯内部" in source:
        return "内部材料"
    return "参考材料"


def workbook_to_items(source: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = find_header_row(sheet)
    headers = [cell_text(cell.value) for cell in sheet[header_row]]
    index = normalize_headers(headers)

    items: list[dict[str, object]] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        row_values = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
        if not any(cell_text(value) for value in row_values):
            continue

        def get_value(key: str) -> str:
            return cell_text(row_values[index[key]]) if key in index else ""

        role = get_value("岗位/场景") or get_value("岗位/职能")
        title = get_value("材料标题")
        summary = get_value("材料简介")
        source_name = get_value("来源")
        link_label = get_value("跳转链接")
        status = get_value("状态")

        link_url = ""
        if "跳转链接" in index:
            link_cell = sheet.cell(row_number, index["跳转链接"] + 1)
            if link_cell.hyperlink and link_cell.hyperlink.target:
                link_url = cell_text(link_cell.hyperlink.target)
            elif link_label.startswith("http://") or link_label.startswith("https://"):
                link_url = link_label

        if status and status != "显示":
            continue

        order_text = get_value("排序")
        order = int(order_text) if order_text.isdigit() else len(items) + 1
        final_summary = summary or infer_summary(role, title, source_name)

        items.append({
            "id": f"material-{order}",
            "order": order,
            "role": role,
            "dimension": classify_dimension(role),
            "title": title,
            "summary": final_summary,
            "linkUrl": link_url,
            "source": source_name,
            "materialType": infer_material_type(link_url, source_name),
            "slug": slugify(title),
        })

    return items


def build_payload(source: Path) -> dict[str, object]:
    items = workbook_to_items(source)
    roles: list[str] = []
    scenes: list[str] = []
    seen_roles: set[str] = set()
    seen_scenes: set[str] = set()
    for item in items:
        label = item["role"]
        if not label:
            continue
        if item["dimension"] == "role":
            if label not in seen_roles:
                roles.append(label)
                seen_roles.add(label)
        else:
            if label not in seen_scenes:
                scenes.append(label)
                seen_scenes.add(label)

    return {
        "meta": {
            "title": "参考材料库",
            "sourceFile": source.name,
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "totalMaterials": len(items),
            "roles": roles,
            "scenes": scenes,
        },
        "items": items,
    }


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python sync_reference_materials_from_excel.py <source.xlsx> <output.js>")
        return 1

    source = Path(sys.argv[1]).expanduser().resolve()
    output = Path(sys.argv[2]).expanduser().resolve()

    payload = build_payload(source)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        "window.REFERENCE_MATERIALS_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
