from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slugify(text: str) -> str:
    safe = []
    for ch in text.lower():
        if ch.isalnum():
            safe.append(ch)
        elif ch in {" ", "-", "_", "/"}:
            safe.append("-")
    slug = "".join(safe).strip("-")
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug or "item"


def find_header_row(sheet) -> int:
    for row_index in range(1, min(sheet.max_row, 10) + 1):
        values = [cell_text(sheet.cell(row_index, col).value) for col in range(1, sheet.max_column + 1)]
        joined = "|".join(values)
        if "Skill名称" in joined or "skill名称" in joined:
            return row_index
    raise ValueError("Could not find header row")


def normalize_headers(headers: list[str]) -> dict[str, int]:
    return {header: index for index, header in enumerate(headers)}


def classify_link(label: str, url: str) -> str:
    text = (label or url).lower()
    if "feishu" in url or "wiki" in text or "教程" in text:
        return "教程"
    if ".zip" in text:
        return "下载包"
    if "skill" in text:
        return "Skill"
    return "链接"


def workbook_to_items(source: Path) -> list[dict[str, object]]:
    workbook = load_workbook(source)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = find_header_row(sheet)
    headers = [cell_text(cell.value) for cell in sheet[header_row]]
    index = normalize_headers(headers)

    items: list[dict[str, object]] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        row_values = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
        if not any(cell_text(value) for value in row_values):
            continue

        get_value = lambda key: cell_text(row_values[index[key]]) if key in index else ""

        name = get_value("Skill名称") or get_value("skill名称")
        category = get_value("Skill类型") or get_value("类型")
        summary = get_value("Skill能力描述") or get_value("skill能力描述")
        link_label = get_value("下载/教程链接")

        link_url = ""
        if "下载/教程链接" in index:
            link_cell = sheet.cell(row_number, index["下载/教程链接"] + 1)
            if link_cell.hyperlink and link_cell.hyperlink.target:
                link_url = cell_text(link_cell.hyperlink.target)
            elif link_label.startswith("http://") or link_label.startswith("https://"):
                link_url = link_label

        order_text = get_value("序号")
        order = int(order_text) if order_text.isdigit() else len(items) + 1

        item = {
            "id": f"skill-{order}",
            "order": order,
            "name": name,
            "summary": summary,
            "owner": get_value("填写人"),
            "package": get_value("所属Skill包"),
            "roles": get_value("目标岗位/职能"),
            "category": category,
            "mainstreamPriority": get_value("主流媒体优先级"),
            "emergingPriority": get_value("新兴媒体优先级"),
            "downloadUrl": link_url,
            "downloadLabel": link_label,
            "linkKind": classify_link(link_label, link_url),
            "slug": slugify(name),
        }
        items.append(item)

    return items


def build_payload(source: Path) -> dict[str, object]:
    items = workbook_to_items(source)
    categories = sorted({item["category"] for item in items})
    packages = sorted({item["package"] for item in items})

    return {
        "meta": {
            "title": "行业 Skill 资源库",
            "sourceFile": source.name,
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "totalSkills": len(items),
            "categories": categories,
            "packages": packages,
        },
        "items": items,
    }


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python sync_skills_from_excel.py <source.xlsx> <output.js>")
        return 1

    source = Path(sys.argv[1]).expanduser().resolve()
    output = Path(sys.argv[2]).expanduser().resolve()

    payload = build_payload(source)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        "window.SKILL_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
